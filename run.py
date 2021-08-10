from openpyxl import workbook,load_workbook
import openpyxl
from prettytable import PrettyTable

data = load_workbook("d:\\harish\\python\\Family-AI\\Data\\FamilyData.xlsx")
wsheet = data["Sheet1"]
w2sheet = data["Sheet2"]
names = []
Place = []
Native = []
Job = []
for cell in wsheet["A"]:
    if cell.value is not None:
        names.append(cell.value)
for cell in w2sheet["D"]:
  if cell.value is not None:
    Place.append(cell.value)
for cell in w2sheet["E"]:
  if cell.value is not None:
    Native.append(cell.value)
for cell in w2sheet["H"]:
  if cell.value is not None:
    Job.append(cell.value)


def compare(names,sentence):
  sent = sentence.split(" ")
  namelen = len(names)
  sentlen = len(sent)
  sent = list(map(lambda x: x.lower(), sent ))
  names = list(map(lambda x: x.lower(), names ))
  for i in range(0,namelen):
    for j in range(0,sentlen):
      if names[i] in sent[j]:
        return names[i]
        pass
      else:
        continue

def dob_finder(statement):
    name = compare(names,statement)
    if name is not None:
        name = name.capitalize()
        for cell in wsheet["A"]:
            if cell.value is not None:
                if cell.value == name:
                    dob = wsheet.cell(row = cell.row,column = 3).value
                    fname = wsheet.cell(row = cell.row,column = 2).value
                    ans = fname + " " + "was Born on" + " " + dob
                    return ans
    else:
      return "Result Not found"

def pob_finder(statement):
    name = compare(names,statement)
    if name is not None:
        name = name.capitalize()
        for cell in wsheet["A"]:
            if cell.value is not None:
                if cell.value == name:
                    pob = wsheet.cell(row = cell.row,column = 4).value
                    fname = wsheet.cell(row = cell.row,column = 2).value
                    ans = fname + " " + "was Born in" + " " + pob
                    return ans
    else:
      return "Result Not found"


def native_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  native = wsheet.cell(row = cell.row,column = 5).value
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  ans = fname + " " + "was a native of" + " " + native
                  return ans
    else:
      return "Result Not found"

def parent_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  dad = wsheet.cell(row = cell.row,column = 6).value
                  mom = wsheet.cell(row = cell.row,column = 7).value
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  ans = dad + " " + "&" + " " + mom + " " +"were the parents of" + " " + fname
                  return ans
    else:
      return "Result Not found"

def father_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  dad = wsheet.cell(row = cell.row,column = 6).value
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  ans = dad + " " + "was the father of" + " " + fname
                  return ans
    else:
      return "Result Not found"

def mother_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  mom = wsheet.cell(row = cell.row,column = 7).value
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  ans = mom + " " + "was the mother of" + " " + fname
                  return ans
    else:
      return "Result Not found"     

def job_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  job = wsheet.cell(row = cell.row,column = 8).value
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  ans = fname + " " + "was a" + " " + job
                  return ans
    else:
      return "Result Not found"     

def bio_finder(statement):
    name = compare(names,statement)
    if name is not None:
      name = name.capitalize()
      for cell in wsheet["A"]:
          if cell.value is not None:
              if cell.value == name:
                  fname = wsheet.cell(row = cell.row,column = 2).value
                  dob = wsheet.cell(row = cell.row,column = 3).value
                  pob = wsheet.cell(row = cell.row,column = 4).value
                  native = wsheet.cell(row = cell.row,column = 5).value
                  father = wsheet.cell(row = cell.row,column = 6).value
                  mother = wsheet.cell(row = cell.row,column = 7).value
                  job = wsheet.cell(row = cell.row,column = 8).value
                  table = PrettyTable()
                  table.field_names = ["Data","Details"]
                  table.add_rows(
                    [
                      ["Name",fname],
                      ["Date of Birth",dob],
                      ["Place of birth",pob],
                      ["Native",native],
                      ["Father",father],
                      ["Mother",mother],
                      ["Job",job]
                    ]
                  )
                  sp = " "
                  speak = "Bio Data of" + sp + fname + sp + "()"  + fname + sp + "was born on" + sp + dob + sp + "at"+ sp + pob + sp + "()" + fname + sp + "was a native of" + sp + native + sp + "()" + father + " " + "&" + " " + mother + " " +"were the parents of" + " " + fname + sp + "()" + fname + " " + "was a" + " " + job
                  return table, speak
    else:
      return "Result Not found"    

def whopbirth_finder(statement):
  place = compare(Place,statement)
  pname = []
  if place is not None:
    place = place.capitalize()
    for cell in wsheet["D"]:
      if cell.value is not None:
        if cell.value == place:
          pname.append(wsheet.cell(row = cell.row,column = 2).value)
  else:
    pname.append("Results not Found") 
  return pname , place

def whonative_finder(statement):
  native = compare(Native,statement)
  nname = []
  if native is not None:
    native = native.capitalize()
    for cell in wsheet["E"]:
      if cell.value is not None:
        if cell.value == native:
          nname.append(wsheet.cell(row = cell.row,column = 2).value)
  else:
    nname.append("Results not Found") 
  return nname , native


